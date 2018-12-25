from django.http import HttpResponse
import datetime, random, calendar, sys
from datetime import date
from calendar import monthrange
from .forms import gensalary, timeentryform
from .models import *
from duradiff.models import Category, Qualification, Citymaster, Statemaster, Resource, Timesheet, Salary
from django.db.models import Sum
import decimal
from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.forms import modelformset_factory
from django.views.generic import View, TemplateView
from django.views.generic.edit import FormView
from django.shortcuts import redirect, render
import openpyxl

@login_required
def gensal(request):
    if request.method == 'POST':  
      form = gensalary(request.POST or None)
      if form.is_valid():
        #mth = int(request.POST.get('month', ''))
        #yr = int(request.POST.get('year', ''))
        #rid = int(request.POST.get('rid', ''))
        #misc = int(request.POST.get('misc', ''))
        #conv = int(request.POST.get('conv', ''))
        mth = int(form.cleaned_data['month'])
        yr = int(form.cleaned_data['year'])
        rid = int(form.cleaned_data['rid'])
        misc = int(form.cleaned_data['misc'])
        conv = int(form.cleaned_data['conv'])
        currentresource = Resource.objects.get(rid = rid)
        basic = currentresource.basicsalary
        misc = currentresource.splallowance
        name = currentresource.firstname + '  ' + currentresource.lastname
        days=monthrange(yr,mth)[-1]
        absentdays = Timesheet.objects.filter(rid=rid).filter(absent__gte=1).count()
        absencededuction = ((basic)/days)*absentdays
        netbasic = basic - absencededuction
        totalOThrs = Timesheet.objects.filter(rid=rid).filter(theday__year=yr).filter(theday__month=mth).aggregate(Sum('OT'))
        totalovertime = sum(totalOThrs.values())
        totalOTamt = Decimal(((basic)/(26 * 8))) * totalovertime
        totalpaybl = Decimal(misc) + Decimal(conv) + Decimal(netbasic) + totalOTamt
        pfamt = (netbasic * 12)/100
        esiamt = (Decimal(totalpaybl) * Decimal(1.75))/100
        shiftpenalty = sum(Timesheet.objects.filter(rid=rid).filter(theday__year=yr).filter(theday__month=mth).aggregate(Sum('shiftshortfall')).values()) * Decimal((basic)/(26 * 8))
        netpaybl = totalpaybl - Decimal(pfamt) - esiamt - shiftpenalty
        sal_obj = Salary(rid=currentresource, month=mth, year=yr,daysinmonth = days, netbasic = netbasic, totalOThours = totalovertime,totalOTamt=totalOTamt,totalpaybl=totalpaybl, pfamt = pfamt, esiamt=esiamt, miscellaneous = misc, conveyance = conv, shiftpenalty = shiftpenalty, netpaybl = netpaybl,name=name,totalabsent=absentdays,absencededuction=absencededuction)
        sal_obj.save()
    else:
        form = gensalary()
    return render(request,'duradiff/hello.html',{'form':form})

@login_required
def uploadform(request):
    if request.method == 'POST':  
      form = uploadform(request.GET or None)
      if form.is_valid():
        #form = uploadform(request)
        #sal_obj = Salary(rid=currentresource, month=mth, year=yr,daysinmonth = days, netbasic = netbasic, totalOThours = totalovertime,totalOTamt=totalOTamt,totalpaybl=totalpaybl, pfamt = pfamt, esiamt=esiamt, miscellaneous = misc, conveyance = conv, shiftpenalty = shiftpenalty, netpaybl = netpaybl,name=name,totalabsent=absentdays,absencededuction=absencededuction)
        #sal_obj.save()
        return
    else:
        form = uploadform()
    return render(request,'duradiff/upload.html',{'form':form})

class TimeView(FormView):
    template_name = 'duradiff/timesheet.html'

    def get(self, request):
        form = timeentryform()
        return render(request, self.template_name,{'form':form})
    
    def post(self,request):
        form = timeentryform(request.POST or None)
        if form.is_valid():
            srid = form.cleaned_data['rid']
            #Accessing Foreign Key Values
            #When you access a field that’s a ForeignKey, in this case rid of Resource model, you’ll get the related model object from the ModelForm.
            #from that you need to access the relevant field, which here is done by srid.rid
            currentresource = Resource.objects.get(rid = srid.rid)
            ridshifthrs = currentresource.shifthrs 
            # commit=False means the form doesn't save at this time.
            # the value of ridshifthrs is fetched from the Resource Model and assigned
            model_instance = form.save(commit=False)
            model_instance.ridshifthrs = ridshifthrs
            # commit defaults to True which means it normally saves.
            model_instance.save()
        args = {'form':form}
        return render(request, self.template_name,args)


#By default, some of Django’s class-based views like this FormView support just a single form per view and multiple forms are not supported
# So the below code did not work

#class MultipleTimeView(FormView):
#    template_name = 'duradiff/multimesheet.html'
    
#    def get(self, request):
#            TimeAddformset = modelformset_factory(Timesheet,form=timeentryform, extra=15)
#            formset = TimeAddformset(queryset=Timesheet.objects.none())
#            args = {'formset':formset}
#            return render(request, 'duradiff/multimesheet.html',args)
    
#    def post(self,request):
#        TimeAddformset = modelformset_factory(Timesheet, form=timeentryform, extra=15)
#        formset = TimeAddformset(request.POST or None)
#        for form in formset.forms:
#          if form.is_valid():
#            srid = form.cleaned_data['rid']
#            currentresource = Resource.objects.get(rid = srid.rid)
#            ridshifthrs = currentresource.shifthrs 
#            model_instance = form.save(commit=False)
#            model_instance.ridshifthrs = ridshifthrs
#            model_instance.save()
#          args = {'form':form}
#        return render(request, 'duradiff/multimesheet.html',args)

#@login_required
#def timesheet(request):
#    if request.method == 'POST':  
#      form = timeentry(request.POST or None)
#      if form.is_valid():
#        #mth = int(request.POST.get('month', ''))
#        #yr = int(request.POST.get('year', ''))
#        #rid = int(request.POST.get('rid', ''))
#        #misc = int(request.POST.get('misc', ''))
#        #conv = int(request.POST.get('conv', ''))
#        theday = form.cleaned_data['theday']
#        timeinhr = form.cleaned_data['timeinhr']
#        rid = int(form.cleaned_data['rid'])
#        timeouthr = form.cleaned_data['timeouthr']
#        absent = form.cleaned_data['absent']
#        fullOTday = form.cleaned_data['fullOTday']
#        timesheet_obj = Timesheet(rid=rid,theday = theday,timeinhr=timeinr, timeouthr=timeouthr,absent=absent,fullOTday=fullOTday)
#        timesheet_obj.save()
#    else:
#        form = timeentry()
#    return render(request,'duradiff/timesheet.html',{'form':form})

from django.http import HttpResponseRedirect
#from .forms import UploadFileForm

# Imaginary function to handle an uploaded file.
#from somewhere import handle_uploaded_file

#def upload_file(request):
  #  if request.method == 'POST' and request.FILES['file']:
     #   form = UploadFileForm(request.POST, request.FILES)
     #   paname = request.FILES['file'].name
      #  #if form.is_valid():
       #     #handle_uploaded_file(request.FILES['file'])
      #      #return HttpResponseRedirect('/success/url/')
  #  else:
     #   form = UploadFileForm()
           
   # return render(request, 'duradiff/upload.html', {'form': form})
@login_required
def upload_file(request):
    if request.method == 'GET':
        return render(request, 'duradiff/upload.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["attendance"]
        #print(worksheet)
        max_row=worksheet.max_row
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)

        return render(request, 'duradiff/upload.html', {"excel_data":excel_data})